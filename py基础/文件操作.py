"""
    @ 文件操作
        1.创建和打开文件
            file = open(filename[,mode[,buffering]])
                file:   被创建的文件对象
                filename:   文件名，需要使用单引号或双引号括起来
                mode:   可选参数，用于指定文件的打开模式，默认为只读（r）
                    r:  只读
                    rb: 以二进制格式打开文件，并采用只读模式。
                    r+: 打开文件后，可以读取文件内容，也可以写入新的内容覆盖原有内容
                    rb+：    以二进制格式打开文件，并且采用读写模式。一般用于非文本文件，如图片，声音等
                    w:  以只写模式打开文件
                    wb: 以二进制格式打开，并且采用只写模式，一般用于非文本文件
                    w+: 打开文件后，先情况原有内容，使其变为一个空的文件，对这个空文件有读写权限
                    wb+:    以二进制格式打开文件，并采用读写模式。
                    a:  以追加模式打开一个文件，用于写入
                    ab: 以二进制格式打开文件，并采用追加模式，用于写入
                    a+: 以读写模式打开文件，并采用追加模式，用于读写
                    ab+:  以二进制格式打开文件，并采用追加模式，用于读写
                buffering:  可选参数，用于指定读写文件的缓冲模式，值为0表示不缓存，值为1表示缓存，值大于1则表示缓冲区的大小

            open()函数打开文件时，默认采用GBK编码，可以通过添加encoding=''参数指定编码

        2.关闭文件
            file.close()

        3.使用with语句：保证with语句执行完毕后关闭已经打开的文件
            with expression as target:
                with-body

                expression: 用于指定一个表达式，这里可以是打开文件的open()函数
                target: 指定一个变量，并且将expression的结果保存到该变量中
                with-body:  用于指定with语句体，其中可以是执行with语句后相关的一些操作语句，如果不想执行任何语句，可以直接使用pass

        4.写入文件内容
            file.write(string)

        5.读取文件
            file.read([size])
                size:   可选参数，指定读取的字符个数，如果省略，则读取所有内容

            seek()方法：read(size)方法读取文件是从头开始的。使用seek()方法可将文件的指针移动到新位置，然后用read(size)方法读取
                file.seek(offset[,whence]
                    offset: 指定移动的字符个数
                    whence: 用于指定从什么位置开始计算。值为0表示从文件头开始计算，值为1表示从当前位置开始计算，值为3便是从文件尾开始计算，默认为0

        6.读取一行
            file.readline()

        7.读取全部行
            file.readlines()    # 返回的是字符串列表，没个元素为一行内容。

    @ 目录操作
        。。。。
"""
import openpyxl

wb = openpyxl.load_workbook("质量控制部.xlsx")
print(wb.sheetnames)
# wb.create_sheet('mySheet')
ws = wb['Sheet1']
# ws["C9"] = "jerry"
# ws.append(["jack", "120", "20"])
# wb.save("质量控制部.xlsx")
for row in ws.iter_rows(min_row=1, max_row=215,max_col=15):
    # print(row)
    for cell in row:
        print(cell.value, end="\t")
    print()
wb.close()
